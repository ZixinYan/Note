import requests  
from bs4 import BeautifulSoup  
import socket  
import openpyxl  
  
def get_website_title(url):  
    try:  
        response = requests.get(url)  
        response.raise_for_status()  # 如果请求失败则引发HTTPError异常  
        soup = BeautifulSoup(response.text, 'html.parser')  
        return soup.title.string  
    except requests.RequestException as e:  
        print(f"Error fetching URL: {e}")  
        return None  
    except (AttributeError, KeyError) as e:  
        print(f"No title tag found: {e}")  
        return None  
  
def check_port(ip, port):  
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)  
    sock.settimeout(1)  # 设置超时时间  
    result = sock.connect_ex((ip, port))  # 尝试连接，返回值是连接状态码  
    if result == 0:  # 0表示连接成功  
        return True  
    return False  
  
# 示例用法  
workbook = openpyxl.Workbook()  # 创建一个新的工作簿  
sheet = workbook.active  # 获取活动工作表  
sheet.title = "Output Results"  # 设置工作表的标题  
  
a = 1  # 从1开始迭代，避免与Excel列的默认索引冲突  
for url in range(1, 256):  # 修改了循环范围以避免无限循环  
    url = f'http://202.122.38.{url}'  
    print(url)  
    if check_port(f"202.122.38.{url}", 80) or check_port(f"202.122.38.{url}", 443):  # 检查特定端口是否开放，例如HTTP的80端口  
        title = get_website_title(url)  
        if title is not None:  
            sheet.cell(row=a, column=1, value=url)  # 将URL写入第一列  
            sheet.cell(row=a, column=2, value=title)  # 将标题写入第二列  
            a += 1  # 增加行索引以继续写入下一个结果  
  
workbook.save("output.xlsx")  # 将工作簿保存为Excel文件